#!/usr/bin/env python3
"""
scripts/export-architecture-xlsx.py — 从 miniprogram-architecture.md 提取表格导出为 xlsx
用法：python3 scripts/export-architecture-xlsx.py
输出：docs/checklists/miniprogram-architecture-tables.xlsx
"""
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

REPO = Path(__file__).parent.parent
MD = REPO / 'docs/checklists/miniprogram-architecture.md'
XLSX = REPO / 'docs/checklists/小程序架构表格.xlsx'


def parse_md_table(table_text):
    """Parse markdown table into list of rows."""
    rows = []
    for line in table_text.strip().split('\n'):
        line = line.strip()
        if line.startswith('|') and line.endswith('|'):
            cells = [c.strip() for c in line[1:-1].split('|')]
            # skip separator line (---|---|---)
            if all(set(c) <= set('-: ') for c in cells):
                continue
            # strip inline code backticks + html br
            cells = [c.replace('`', '').replace('<br/>', ' / ').replace('<br>', ' / ') for c in cells]
            rows.append(cells)
    return rows


def extract_section_tables(md_text, section_header):
    """Find tables under a given ## section."""
    lines = md_text.split('\n')
    in_section = False
    tables = []
    current = []

    for line in lines:
        if line.startswith('## '):
            if in_section:
                if current:
                    tables.append(current)
                    current = []
            in_section = (line == section_header)
            continue
        if in_section and line.startswith('|') and not all(set(c.strip()) <= set('-: ') for c in line[1:-1].split('|') if c.strip()):
            current.append(line)
        elif in_section and not line.startswith('|') and current:
            tables.append(current)
            current = []

    if in_section and current:
        tables.append(current)
    return tables


def write_sheet(ws, title, tables):
    """Write workbook sheet with multiple tables."""
    # Title
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='left')

    row = 3
    for i, table_lines in enumerate(tables):
        if not table_lines:
            continue
        if len(tables) > 1:
            ws.cell(row=row, column=1, value=f'【表 {i+1}】').font = Font(bold=True, italic=True)
            row += 1
        rows = parse_md_table('\n'.join(table_lines))
        if not rows:
            continue
        for r in rows:
            for col_idx, val in enumerate(r, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                if row == (row - len(r) + (1 if len(tables) > 1 else 0)) and r == rows[0]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill('solid', fgColor='D0E8FF')
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            row += 1
        row += 1

    # auto column width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                # 中文宽度 2，其他 1
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def main():
    md_text = MD.read_text(encoding='utf-8')

    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    # §2 前端功能矩阵
    ws2 = wb.create_sheet('§2 前端功能矩阵')
    tables = extract_section_tables(md_text, '## §2 前端功能矩阵（29 页 × 功能 × 后端 API）')
    write_sheet(ws2, '§2 前端功能矩阵 — 29 页 × 功能 × 后端 API × V0.2.X', tables)

    # §3 组件复用矩阵
    ws3 = wb.create_sheet('§3 组件复用矩阵')
    tables = extract_section_tables(md_text, '## §3 组件复用矩阵（17 个 components）')
    write_sheet(ws3, '§3 组件复用矩阵 — 17 components × props × 事件 × 使用页面', tables)

    # §6 后端 API 分布
    ws6 = wb.create_sheet('§6 后端 API 分布')
    tables = extract_section_tables(md_text, '## §6 后端 API 分布（按 module 排序）')
    write_sheet(ws6, '§6 后端 API 分布 — 33 module × 250+ action × V0.2.X 增量', tables)

    # §0 元信息（README-style）
    ws0 = wb.create_sheet('README', 0)  # 第一个
    ws0['A1'] = 'MiniProgram Architecture — 表格汇总'
    ws0['A1'].font = Font(size=16, bold=True)
    ws0['A3'] = '生成时间'
    ws0['B3'] = '2026-07-26 17:53 CST'
    ws0['A4'] = '数据源'
    ws0['B4'] = '/zcf:init-project #20 init-architect 校准 + 13 commits 沉淀 + 当前 app.json/routes.ts 扫描'
    ws0['A5'] = 'Sheets'
    ws0['B5'] = '§2 前端功能矩阵 / §3 组件复用矩阵 / §6 后端 API 分布'
    ws0['A6'] = '关联文件'
    ws0['B6'] = 'docs/checklists/miniprogram-architecture.md + miniprogram-architecture.pdf'
    ws0['A7'] = '维护者'
    ws0['B7'] = '/zcf:workflow 自动生成；V0.3.1 增量时同步更新'

    for row in [3, 4, 5, 6, 7]:
        ws0.cell(row=row, column=1).font = Font(bold=True)

    for col in ws0.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_len = max(max_len, length)
        ws0.column_dimensions[col_letter].width = min(max_len + 2, 80)

    wb.save(XLSX)
    print(f'✅ xlsx 生成: {XLSX}')
    print(f'   - Sheets: README + §2/§3/§6 = 4 sheets')
    print(f'   - §2 表: {len(extract_section_tables(md_text, "## §2 前端功能矩阵（29 页 × 功能 × 后端 API）"))} 个')
    print(f'   - §3 表: {len(extract_section_tables(md_text, "## §3 组件复用矩阵（17 个 components）"))} 个')
    print(f'   - §6 表: {len(extract_section_tables(md_text, "## §6 后端 API 分布（按 module 排序）"))} 个')


if __name__ == '__main__':
    main()